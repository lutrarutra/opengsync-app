import os
import io
from typing import Literal

import pandas as pd

import openpyxl
from openpyxl import styles as openpyxl_styles
from openpyxl.utils import get_column_letter

from flask import Blueprint, render_template, request, flash, url_for, Response
from flask_htmx import make_response

from opengsync_db import models, PAGE_LIMIT
from opengsync_db.categories import PoolStatus, LibraryStatus, PrepStatus, SeqRequestStatus, LibraryType, LabChecklistType

from ... import db, forms, logger, logic
from ...core import wrappers, exceptions
from ...core.RunTime import runtime
from ...tools.spread_sheet_components import TextColumn
from ...tools import StaticSpreadSheet


lab_preps_htmx = Blueprint("lab_preps_htmx", __name__, url_prefix="/htmx/lab_preps/")


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get(current_user: models.User):
    context = logic.lab_prep.get_table_context(current_user=current_user, request=request)
    return make_response(render_template(**context))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def create(current_user: models.User):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if request.method == "GET":
        return forms.models.LabPrepForm(form_type="create").make_response()
    
    form = forms.models.LabPrepForm(formdata=request.form, form_type="create")
    return form.process_request(current_user)


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def edit(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if request.method == "GET":
        return forms.models.LabPrepForm( form_type="edit", lab_prep=lab_prep).make_response()
        
    return forms.models.LabPrepForm(formdata=request.form, form_type="edit", lab_prep=lab_prep).process_request(current_user)


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def complete(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    for pool in lab_prep.pools:
        if pool.status < PoolStatus.STORED:
            pool.status = PoolStatus.STORED
            db.pools.update(pool)

    for library in lab_prep.libraries:
        is_prepared = True
        for sr_library in library.seq_request.libraries:
            is_prepared = sr_library.status >= LibraryStatus.POOLED and is_prepared
            if not is_prepared:
                break
        if is_prepared:
            library.seq_request.status = SeqRequestStatus.PREPARED
            db.libraries.update(library)

    lab_prep.status = PrepStatus.COMPLETED
    db.lab_preps.update(lab_prep)

    flash("Lab prep completed!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["DELETE"])
def delete(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if lab_prep.status != PrepStatus.PREPARING:
        flash("Cannot delete completed prep.", "warning")
        return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))
    
    db.lab_preps.delete(lab_prep_id=lab_prep.id)
    flash("Lab prep deleted!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_preps"))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def uncomplete(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()

    lab_prep.status = PrepStatus.PREPARING
    db.lab_preps.update(lab_prep)

    flash("Lab prep completed!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["DELETE"], arg_params=["library_id"])
def remove_library(current_user: models.User, lab_prep_id: int, library_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if lab_prep.status != PrepStatus.PREPARING:
        raise exceptions.BadRequestException()
    
    db.lab_preps.remove_library(lab_prep_id=lab_prep.id, library_id=library_id)

    flash("Library Removed!.", "success")
    context = logic.library.get_table_context(current_user=current_user, request=request, lab_prep=lab_prep)
    return make_response(render_template(**context))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET"], arg_params=["checklist_id"])
def download_template(current_user: models.User, lab_prep_id: int, direction: Literal["rows", "columns"], checklist_id: int | None = None) -> Response:
    if direction not in ("rows", "columns"):
        raise exceptions.BadRequestException()
    
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if checklist_id is not None:
        try:
            checklist = LabChecklistType.get(checklist_id)
        except ValueError:
            raise exceptions.BadRequestException(f"Unknown checklist ID: {checklist_id}")
    else:
        checklist = lab_prep.checklist_type

    if runtime.app.static_folder is None:
        logger.error("Static folder not set")
        raise ValueError("Static folder not set")

    filepath = os.path.join(runtime.app.static_folder, "resources", "templates", "library_prep", checklist.prep_file_name)

    if not os.path.exists(filepath):
        logger.error(f"File not found: {filepath}")
        raise exceptions.NotFoundException()

    template = openpyxl.load_workbook(filepath)

    n = 12 if direction == "rows" else 8
    pattern = [True] * n + [False] * n

    def if_color(i):
        return pattern[i]

    prep_table = template["prep_table"]
    column_mapping: dict[str, str] = {}
    
    for col_i in range(0, min(prep_table.max_column, 96)):
        col = get_column_letter(col_i + 1)
        column_name = prep_table[f"{col}1"].value
        column_mapping[column_name] = col
        
        for row_idx, cell in enumerate(prep_table[col][1:]):
            if if_color(row_idx % (n * 2)):
                cell.fill = openpyxl_styles.PatternFill(start_color="ced4da", end_color="ced4da", fill_type="solid")
            else:
                cell.fill = openpyxl_styles.PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

    for row_idx, cell in enumerate(prep_table[column_mapping["plate_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")

    for row_idx, cell in enumerate(prep_table[column_mapping["index_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")
        
    for i, library in enumerate(lab_prep.libraries):
        library_id_cell = prep_table[f"{column_mapping['library_id']}{i + 2}"]
        library_name_cell = prep_table[f"{column_mapping['library_name']}{i + 2}"]
        requestor_cell = prep_table[f"{column_mapping['requestor']}{i + 2}"]
        sequence_i7_cell = prep_table[f"{column_mapping['sequence_i7']}{i + 2}"]
        sequence_i5_cell = prep_table[f"{column_mapping['sequence_i5']}{i + 2}"]
        # kit_i7_cell = active_sheet[f"{column_mapping['kit_i7']}{i + 2}"]
        # kit_i5_cell = active_sheet[f"{column_mapping['kit_i5']}{i + 2}"]
        name_i7_cell = prep_table[f"{column_mapping['name_i7']}{i + 2}"]
        name_i5_cell = prep_table[f"{column_mapping['name_i5']}{i + 2}"]
        library_id_cell.value = library.id
        library_name_cell.value = library.name
        requestor_cell.value = library.seq_request.requestor.name
        if len(library.indices) > 0:
            name_i7_cell.value = library.indices[0].name_i7
            name_i5_cell.value = library.indices[0].name_i5
        sequence_i7_cell.value = library.sequences_i7_str(";")
        sequence_i5_cell.value = library.sequences_i5_str(";")

    if "FLEX_table" in template.sheetnames:
        flex_table = template["FLEX_table"]
        column_mapping: dict[str, str] = {}
    
        for col_i in range(0, min(flex_table.max_column, 96)):
            col = get_column_letter(col_i + 1)
            column_name = flex_table[f"{col}1"].value
            column_mapping[column_name] = col

        i = 2
        for library in lab_prep.libraries:
            if library.type == LibraryType.TENX_SC_GEX_FLEX:
                for sample_link in library.sample_links:
                    sample_num_cell = flex_table[f"{column_mapping['sample_num']}{i}"]
                    sample_name_cell = flex_table[f"{column_mapping['sample_name']}{i}"]
                    sample_num_cell.value = i - 1
                    sample_name_cell.value = sample_link.sample.name
                    i += 1

    if "10X_table" in template.sheetnames:
        tenx_table = template["10X_table"]
        column_mapping: dict[str, str] = {}
    
        for col_i in range(0, min(tenx_table.max_column, 96)):
            col = get_column_letter(col_i + 1)
            column_name = tenx_table[f"{col}1"].value
            column_mapping[column_name] = col

        i = 2
        for library in lab_prep.libraries:
            if library.type in (LibraryType.TENX_SC_GEX_3PRIME, LibraryType.TENX_SC_GEX_5PRIME):
                for sample_link in library.sample_links:
                    sample_num_cell = tenx_table[f"{column_mapping['sample_num']}{i}"]
                    sample_name_cell = tenx_table[f"{column_mapping['sample_name']}{i}"]
                    library_name_cell = tenx_table[f"{column_mapping['library_name']}{i}"]
                    library_id_cell = tenx_table[f"{column_mapping['library_id']}{i}"]
                    requestor_cell = tenx_table[f"{column_mapping['requestor']}{i}"]

                    sample_num_cell.value = i - 1
                    sample_name_cell.value = sample_link.sample.name
                    library_name_cell.value = library.name
                    library_id_cell.value = library.id
                    requestor_cell.value = library.seq_request.requestor.name
                    i += 1
            
    bytes_io = io.BytesIO()
    template.save(bytes_io)

    bytes_io.seek(0)
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
    return Response(
        bytes_io, mimetype=mimetype,
        headers={"Content-disposition": f"attachment; filename={lab_prep.name}_{lab_prep.checklist_type.abbreviation}_{direction}.xlsx"}
    )


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def prep_table_upload_form(current_user: models.User, lab_prep_id: int) -> Response:
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if request.method == "GET":
        form = forms.workflows.library_prep.LibraryPrepForm(lab_prep=lab_prep)
        return form.make_response()
    else:
        form = forms.workflows.library_prep.LibraryPrepForm(lab_prep=lab_prep, formdata=request.form | request.files)
        return form.process_request(user=current_user)
    

@wrappers.htmx_route(lab_preps_htmx, db=db)
def checklist(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    checklist = lab_prep.get_checklist()
    return make_response(
        render_template(
            "components/checklists/lab_prep.html",
            lab_prep=lab_prep, **checklist
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get_files(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()

    return make_response(
        render_template(
            "components/file-list.html",
            files=lab_prep.media_files, delete="lab_preps_htmx.delete_file",
            delete_context={"lab_prep_id": lab_prep_id}
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["DELETE"])
def delete_file(current_user: models.User, lab_prep_id: int, file_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if (file := db.media_files.get(file_id)) is None:
        raise exceptions.NotFoundException()
    
    if file not in lab_prep.media_files:
        raise exceptions.BadRequestException()

    file_path = os.path.join(runtime.app.media_folder, file.path)
    if os.path.exists(file_path):
        os.remove(file_path)
    db.media_files.delete(file_id=file.id)

    logger.info(f"Deleted file '{file.name}' from prep (id='{lab_prep_id}')")
    flash(f"Deleted file '{file.name}' from prep.", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def file_form(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if request.method == "GET":
        form = forms.file.LabPrepAttachmentForm(lab_prep=lab_prep)
        return form.make_response()
    elif request.method == "POST":
        form = forms.file.LabPrepAttachmentForm(lab_prep=lab_prep, formdata=request.form | request.files)
        return form.process_request(current_user)
    else:
        raise exceptions.MethodNotAllowedException()
    

@wrappers.htmx_route(lab_preps_htmx, db=db)
def plates_tab(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    return make_response(render_template("components/plates.html", plates=lab_prep.plates, from_page=f"lab_prep@{lab_prep.id}"))
    

@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def comment_form(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if request.method == "GET":
        form = forms.comment.LabPrepCommentForm(lab_prep=lab_prep)
        return form.make_response()
    elif request.method == "POST":
        form = forms.comment.LabPrepCommentForm(lab_prep=lab_prep, formdata=request.form)
        return form.process_request(current_user)
    else:
        raise exceptions.MethodNotAllowedException()


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get_comments(current_user: models.User, lab_prep_id: int):
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()

    return make_response(
        render_template(
            "components/comment-list.html",
            comments=lab_prep.comments
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get_mux_table(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    df = db.pd.get_lab_prep_pooling_table(lab_prep.id)

    df = df.sort_values(by=["library_name", "sample_pool", "sample_name"])

    mux_table = {
        "sample_name": [],
        "library_name": [],
        "sample_pool": [],
        "barcode": [],
        "pattern": [],
        "read": []
    }

    for _, row in df.iterrows():
        if row["mux_type_id"] is None:
            continue
        
        mux_table["sample_name"].append(row["sample_name"])
        mux_table["library_name"].append(row["library_name"])
        mux_table["sample_pool"].append(row["sample_pool"])
        if (mux := row.get("mux")) is None:
            mux_table["barcode"].append(None)
            mux_table["pattern"].append(None)
            mux_table["read"].append(None)
        else:
            mux_table["barcode"].append(mux.get("barcode"))
            mux_table["pattern"].append(mux.get("pattern"))
            mux_table["read"].append(mux.get("read"))

    df = pd.DataFrame(mux_table)
    if df["pattern"].isna().all():
        df = df.drop(columns=["pattern"])
    if df["read"].isna().all():
        df = df.drop(columns=["read"])
        
    columns = []
    for col in df.columns:
        columns.append(
            TextColumn(
                col,
                col.replace("_", " ").title().replace("Id", "ID").replace("Cmo", "CMO"),
                {
                    "sample_name": 250,
                    "library_name": 300,
                    "sample_pool": 200,
                    "barcode": 100,
                    "read": 80,
                    "pattern": 150
                }.get(col, 100),
                max_length=1000
            )
        )

    spreadsheet = StaticSpreadSheet(df, columns=columns, id=f"lab_prep_mux_table-{lab_prep_id}")

    return make_response(render_template("components/itable.html", spreadsheet=spreadsheet))


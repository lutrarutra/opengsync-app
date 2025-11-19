import os

import pandas as pd
import openpyxl

from opengsync_db.categories import LabChecklistType

common_columns = [
    "library_id",
    "library_name",
    "requestor",
    "pool",
    "plate_well",
    "index_well",
    "kit_i7",
    "name_i7",
    "sequence_i7",
    "kit_i5",
    "name_i5",
    "sequence_i5",
    "lib_conc_ng_ul",
]

prep_table_template_dir = "prep_tables"


def common_prep_table_test(table_path: str):
    assert os.path.exists(table_path)

    wb = openpyxl.load_workbook(table_path)
    assert wb.sheetnames[0] == "prep_table"

    df = pd.read_excel(table_path, "prep_table")

    assert len(df) == 96

    for col in common_columns:
        assert col in df.columns, f"Column not found: {col}"
        assert df[col].isna().sum() == len(df), f"Not empty: {col}"

    return wb


def test_default_table():
    wb = common_prep_table_test(os.path.join(prep_table_template_dir, LabChecklistType.CUSTOM.prep_file_name))


def test_rna_table():
    wb = common_prep_table_test(os.path.join(prep_table_template_dir, LabChecklistType.RNA_SEQ.prep_file_name))


def test_wgs_table():
    wb = common_prep_table_test(os.path.join(prep_table_template_dir, LabChecklistType.WGS.prep_file_name))


def test_qseq_table():
    wb = common_prep_table_test(os.path.join(prep_table_template_dir, LabChecklistType.QUANT_SEQ.prep_file_name))


def test_tenx_table():
    wb = common_prep_table_test(os.path.join(prep_table_template_dir, LabChecklistType.TENX.prep_file_name))
    assert "10X_table" in wb.sheetnames
    tenx_table = pd.read_excel(os.path.join(prep_table_template_dir, LabChecklistType.TENX.prep_file_name), "10X_table")
    for col in ["sample_num", "sample_name"]:
        assert col in tenx_table.columns, f"Column not found: {col}"
    assert "FLEX_table" in wb.sheetnames
    flex_table = pd.read_excel(os.path.join(prep_table_template_dir, LabChecklistType.TENX.prep_file_name), "FLEX_table")
    for col in ["sample_num", "sample_name"]:
        assert col in flex_table.columns, f"Column not found: {col}"


def test_smartseq_table():
    wb = common_prep_table_test(os.path.join(prep_table_template_dir, LabChecklistType.SMART_SEQ.prep_file_name))


def test_wes_table():
    wb = common_prep_table_test(os.path.join(prep_table_template_dir, LabChecklistType.WES.prep_file_name))
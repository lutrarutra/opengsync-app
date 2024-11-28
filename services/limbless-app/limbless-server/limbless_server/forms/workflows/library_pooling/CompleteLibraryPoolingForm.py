import os
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from flask import Response, url_for, flash, current_app
from flask_htmx import make_response

from limbless_db import models
from limbless_db.categories import PoolType, SeqRequestStatus, LibraryStatus

from .... import logger, db, tools
from ...MultiStepForm import MultiStepForm


class CompleteLibraryPoolingForm(MultiStepForm):
    _template_path = "workflows/library_pooling/complete-pooling.html"
    _workflow_name = "library_pooling"
    _step_name = "complete_library_pooling"

    def __init__(self, uuid: str | None, previous_form: Optional[MultiStepForm] = None, formdata: dict = {}):
        MultiStepForm.__init__(
            self, workflow=CompleteLibraryPoolingForm._workflow_name,
            step_name=CompleteLibraryPoolingForm._step_name, uuid=uuid,
            formdata=formdata, previous_form=previous_form, step_args={}
        )

    def prepare(self):
        barcode_table = self.tables["barcode_table"]
        barcode_table = tools.check_indices(barcode_table, groupby="pool")
        self._context["barcode_table"] = barcode_table

    def validate(self) -> bool:
        validated = super().validate()
        return validated

    def process_request(self, user: models.User) -> Response:
        if not self.validate():
            return self.make_response()

        lab_prep_id = self.metadata["lab_prep_id"]
        if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
            logger.error(f"{self.uuid}: LabPrep not found")
            raise ValueError(f"{self.uuid}: LabPrep not found")

        for library in lab_prep.libraries:
            library.pool_id = None
            library.status = LibraryStatus.FAILED
            library = db.update_library(library)
            library = db.remove_library_indices(library_id=library.id)
        
        barcode_table = self.tables["barcode_table"]
        barcode_table.loc[barcode_table["name_i7"].notna(), "name_i7"] = barcode_table.loc[barcode_table["name_i7"].notna(), "name_i7"].astype(str)
        barcode_table.loc[barcode_table["name_i5"].notna(), "name_i5"] = barcode_table.loc[barcode_table["name_i5"].notna(), "name_i5"].astype(str)
        library_table = self.tables["library_table"]
        
        if (lab_prep_id := self.metadata.get("lab_prep_id")) is None:
            logger.error(f"{self.uuid}: LabPrep id not found")
            raise ValueError(f"{self.uuid}: LabPrep id not found")
        
        if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
            logger.error(f"{self.uuid}: LabPrep not found")
            raise ValueError(f"{self.uuid}: LabPrep not found")
        
        for pool in lab_prep.pools:
            db.delete_pool(pool.id)

        selected_libraries = library_table[~library_table["pool"].astype(str).str.strip().str.lower().isin(["x", "t"])].copy()
        
        pools = {}
        if len(selected_libraries["pool"].unique()) > 1:
            for pool_suffix, _df in selected_libraries.groupby("pool"):
                pools[pool_suffix] = db.create_pool(
                    name=f"{lab_prep.name}_{pool_suffix}", pool_type=PoolType.INTERNAL,
                    contact_email=user.email, contact_name=user.name, owner_id=user.id,
                    lab_prep_id=lab_prep.id
                )
        elif len(selected_libraries) > 0:
            selected_libraries["pool"] = "1"
            library_table["pool"] = "1"
            pools["1"] = db.create_pool(
                name=lab_prep.name, pool_type=PoolType.INTERNAL,
                contact_email=user.email, contact_name=user.name, owner_id=user.id,
                lab_prep_id=lab_prep.id
            )

        request_ids = []
        for i, (idx, row) in enumerate(library_table.iterrows()):
            if (library := db.get_library(row["library_id"])) is None:
                logger.error(f"{self.uuid}: Library {row['library_id']} not found")
                raise ValueError(f"{self.uuid}: Library {row['library_id']} not found")

            if str(row["pool"]).strip().lower() == "t":
                continue
            elif str(row["pool"]).strip().lower() == "x":
                library.pool_id = None
                library = db.update_library(library)
                library = db.remove_library_indices(library_id=library.id)
            else:
                df = barcode_table[barcode_table["library_id"] == row["library_id"]].copy()

                seq_i7s = df["sequence_i7"].values
                seq_i5s = df["sequence_i5"].values
                name_i7s = df["name_i7"].values
                name_i5s = df["name_i5"].values

                kit_i7_id = row["kit_i7_id"] if pd.notna(row["kit_i7_id"]) else None
                kit_i5_id = row["kit_i5_id"] if pd.notna(row["kit_i5_id"]) else None

                for j in range(max(len(seq_i7s), len(seq_i5s))):
                    library = db.add_library_index(
                        library_id=library.id, index_kit_i7_id=kit_i7_id, index_kit_i5_id=kit_i5_id,
                        name_i7=name_i7s[j] if len(name_i7s) > j and pd.notna(name_i7s[j]) else None,
                        name_i5=name_i5s[j] if len(name_i5s) > j and pd.notna(name_i5s[j]) else None,
                        sequence_i7=seq_i7s[j] if len(seq_i7s) > j and pd.notna(seq_i7s[j]) else None,
                        sequence_i5=seq_i5s[j] if len(seq_i5s) > j and pd.notna(seq_i5s[j]) else None,
                    )
                # library_table.at[idx, "name_i7"] = row["name_i7"] if pd.notna(row["name_i7"]) else None
                # library_table.at[idx, "name_i5"] = row["name_i5"] if pd.notna(row["name_i5"]) else None

                library.pool_id = None
                library = db.update_library(library)
                library = db.pool_library(library_id=library.id, pool_id=pools[row["pool"]].id)
                library.status = LibraryStatus.POOLED
                library = db.update_library(library)

                if library.seq_request_id not in request_ids:
                    request_ids.append(library.seq_request_id)
        
        if lab_prep.prep_file is not None:
            wb = openpyxl.load_workbook(os.path.join(current_app.config["MEDIA_FOLDER"], lab_prep.prep_file.path))
            active_sheet = wb["prep_table"]
            
            column_mapping: dict[str, str] = {}
            for col_i in range(1, active_sheet.max_column):
                col = get_column_letter(col_i + 1)
                column_name = active_sheet[f"{col}1"].value
                column_mapping[column_name] = col
            
            for i, (idx, row) in enumerate(library_table.iterrows()):
                if (library := db.get_library(int(row["library_id"]))) is None:
                    logger.error(f"{self.uuid}: Library {row['library_id']} not found")
                    raise ValueError(f"{self.uuid}: Library {row['library_id']} not found")

                active_sheet[f"{column_mapping['sequence_i7']}{i + 2}"].value = library.sequences_i7_str(sep=";")
                active_sheet[f"{column_mapping['sequence_i5']}{i + 2}"].value = library.sequences_i5_str(sep=";")
                active_sheet[f"{column_mapping['name_i7']}{i + 2}"].value = ";".join([index.name_i7 for index in library.indices if index.name_i7])
                active_sheet[f"{column_mapping['name_i5']}{i + 2}"].value = ";".join([index.name_i5 for index in library.indices if index.name_i5])
                active_sheet[f"{column_mapping['pool']}{i + 2}"].value = row["pool"]
                active_sheet[f"{column_mapping['kit_i7']}{i + 2}"].value = row["kit_i7"]
                active_sheet[f"{column_mapping['kit_i5']}{i + 2}"].value = row["kit_i5"]
                active_sheet[f"{column_mapping['index_well']}{i + 2}"].value = row["index_well"]

            logger.debug(f"Overwriting existing file: {os.path.join(current_app.config['MEDIA_FOLDER'], lab_prep.prep_file.path)}")
            wb.save(os.path.join(current_app.config["MEDIA_FOLDER"], lab_prep.prep_file.path))

        for request_id in request_ids:
            if (seq_request := db.get_seq_request(request_id)) is None:
                logger.error(f"{self.uuid}: SeqRequest {request_id} not found")
                raise ValueError(f"{self.uuid}: SeqRequest {request_id} not found")
            
            prepared = True
            for library in seq_request.libraries:
                prepared = prepared and library.status.id >= LibraryStatus.POOLED.id

            if prepared and seq_request.status == SeqRequestStatus.ACCEPTED:
                seq_request.status = SeqRequestStatus.PREPARED
                seq_request = db.update_seq_request(seq_request)
        
        self.complete()
        flash("Library Indexing completed!", "success")
        return make_response(redirect=url_for("lab_preps_page.lab_prep_page", lab_prep_id=lab_prep.id))

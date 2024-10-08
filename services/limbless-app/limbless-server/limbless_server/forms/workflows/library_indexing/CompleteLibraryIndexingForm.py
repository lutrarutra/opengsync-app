import os
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from flask import Response, url_for, flash, current_app
from flask_htmx import make_response

from limbless_db import models
from limbless_db.categories import PoolType

from .... import logger, db, tools
from ...TableDataForm import TableDataForm
from ...HTMXFlaskForm import HTMXFlaskForm


class CompleteLibraryIndexingForm(HTMXFlaskForm, TableDataForm):
    _template_path = "workflows/library_indexing/indexing-3.html"
    _form_label = "library_indexing_form"

    def __init__(self, previous_form: Optional[TableDataForm] = None, formdata: dict = {}, uuid: Optional[str] = None):
        if uuid is None:
            uuid = formdata.get("file_uuid")
        HTMXFlaskForm.__init__(self, formdata=formdata)
        TableDataForm.__init__(self, dirname="library_indexing", uuid=uuid, previous_form=previous_form)

    def prepare(self):
        barcode_table = self.tables["barcode_table"]
        barcode_table = tools.check_indices(barcode_table, groupby="pool")

        self._context["barcode_table"] = barcode_table

    def validate(self) -> bool:
        validated = super().validate()
        return validated

    def process_request(self, user: models.User) -> Response:
        if not self.validate():
            return self.make_response()
        
        barcode_table = self.tables["barcode_table"]
        library_table = self.tables["library_table"]
        
        if (lab_prep_id := self.metadata.get("lab_prep_id")) is None:
            logger.error(f"{self.uuid}: LabPrep id not found")
            raise ValueError(f"{self.uuid}: LabPrep id not found")
        
        if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
            logger.error(f"{self.uuid}: LabPrep not found")
            raise ValueError(f"{self.uuid}: LabPrep not found")
        
        for pool in lab_prep.pools:
            db.delete_pool(pool.id)
        
        pools = {}
        if len(library_table["pool"].unique()) > 1:
            for pool_suffix, _df in library_table.groupby("pool"):
                pools[pool_suffix] = db.create_pool(
                    name=f"{lab_prep.name}_{pool_suffix}", pool_type=PoolType.INTERNAL,
                    contact_email=user.email, contact_name=user.name, owner_id=user.id,
                    lab_prep_id=lab_prep.id
                )
        else:
            library_table["pool"] = 1
            pools[1] = db.create_pool(
                name=lab_prep.name, pool_type=PoolType.INTERNAL,
                contact_email=user.email, contact_name=user.name, owner_id=user.id,
                lab_prep_id=lab_prep.id
            )

        for i, (idx, row) in enumerate(library_table.iterrows()):
            if (library := db.get_library(row["library_id"])) is None:
                logger.error(f"{self.uuid}: Library {row['library_id']} not found")
                raise ValueError(f"{self.uuid}: Library {row['library_id']} not found")
            
            library = db.remove_library_indices(library.id)

            df = barcode_table[barcode_table["library_id"] == row["library_id"]].copy()

            seq_i7s = df["sequence_i7"].values
            seq_i5s = df["sequence_i5"].values

            for j in range(max(len(seq_i7s), len(seq_i5s))):
                library = db.add_library_index(
                    library_id=library.id,
                    name_i7=row["name_i7"] if pd.notna(row["name_i7"]) else None,
                    name_i5=row["name_i5"] if pd.notna(row["name_i5"]) else None,
                    sequence_i7=seq_i7s[j] if len(seq_i7s) > j and pd.notna(seq_i7s[j]) else None,
                    sequence_i5=seq_i5s[j] if len(seq_i5s) > j and pd.notna(seq_i5s[j]) else None,
                )
            library_table.at[idx, "name_i7"] = row["name_i7"] if pd.notna(row["name_i7"]) else None
            library_table.at[idx, "name_i5"] = row["name_i5"] if pd.notna(row["name_i5"]) else None

            library.pool_id = None
            library = db.update_library(library)
            library = db.pool_library(library_id=library.id, pool_id=pools[row["pool"]].id)
        
        if lab_prep.prep_file is not None:
            wb = openpyxl.load_workbook(os.path.join(current_app.config["MEDIA_FOLDER"], lab_prep.prep_file.path))
            active_sheet = wb["prep_table"]
            
            column_mapping: dict[str, str] = {}
            for col_i in range(1, active_sheet.max_column):
                col = get_column_letter(col_i + 1)
                column_name = active_sheet[f"{col}1"].value
                column_mapping[column_name] = col

            logger.debug(column_mapping)
            
            for i, (idx, row) in enumerate(library_table.iterrows()):
                df = barcode_table[barcode_table["library_id"] == row["library_id"]]
                sequence_i7 = ";".join([s for s in df["sequence_i7"].values if pd.notna(s)])
                sequence_i5 = ";".join([s for s in df["sequence_i5"].values if pd.notna(s)])
                name_i7 = ";".join([s for s in df["name_i7"].values if pd.notna(s)])
                name_i5 = ";".join([s for s in df["name_i5"].values if pd.notna(s)])
                active_sheet[f"{column_mapping['sequence_i7']}{i + 2}"].value = sequence_i7
                active_sheet[f"{column_mapping['sequence_i5']}{i + 2}"].value = sequence_i5
                active_sheet[f"{column_mapping['name_i7']}{i + 2}"].value = name_i7
                active_sheet[f"{column_mapping['name_i5']}{i + 2}"].value = name_i5
                active_sheet[f"{column_mapping['pool']}{i + 2}"].value = row["pool"]
                active_sheet[f"{column_mapping['kit_i7']}{i + 2}"].value = row["kit_i7"]
                active_sheet[f"{column_mapping['kit_i5']}{i + 2}"].value = row["kit_i5"]
                active_sheet[f"{column_mapping['index_well']}{i + 2}"].value = row["index_well"]

            logger.debug(f"Overwriting existing file: {os.path.join(current_app.config['MEDIA_FOLDER'], lab_prep.prep_file.path)}")
            wb.save(os.path.join(current_app.config["MEDIA_FOLDER"], lab_prep.prep_file.path))

        flash("Library Indexing completed!", "success")
        return make_response(redirect=url_for("lab_preps_page.lab_prep_page", lab_prep_id=lab_prep.id))

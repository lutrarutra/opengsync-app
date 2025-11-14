import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter

from flask import Response, url_for, flash
from flask_htmx import make_response

from opengsync_db import models
from opengsync_db.categories import FeatureType

from .. import db, logger  # noqa
from ..tools import utils
from ..core import runtime
from ..tools.spread_sheet_components import CategoricalDropDown, DuplicateCellValue, SpreadSheetColumn
from .HTMXFlaskForm import HTMXFlaskForm
from .SpreadsheetInput import SpreadsheetInput


class LibraryProtocolSelectForm(HTMXFlaskForm):
    _template_path = "forms/select_library_protocol.html"

    def __init__(self, lab_prep: models.LabPrep, formdata: dict | None = None):
        super().__init__(formdata=formdata)
        self.lab_prep = lab_prep
        self._context["lab_prep"] = lab_prep

        self.protocol_mappings = {protocol.id: protocol.name for protocol in db.protocols.find(limit=None, sort_by="name")[0]}
        self.library_mappings = {lib.id: lib.name for lib in lab_prep.libraries}

        columns: list = [
            CategoricalDropDown("library_id", "Library", 300, categories=self.library_mappings, required=True),
            CategoricalDropDown("protocol_id", "Protocol", 1000, categories=self.protocol_mappings, required=False),
        ]

        self.spreadsheet: SpreadsheetInput = SpreadsheetInput(
            columns=columns, csrf_token=self._csrf_token,
            post_url=url_for("lab_preps_htmx.protocol_form", lab_prep_id=lab_prep.id),
            formdata=formdata
        )

        template = self.__get_template()
        self.spreadsheet.set_data(template)

    def __get_template(self):
        data = {
            "library_id": [],
            "protocol_id": [],
        }

        for library in self.lab_prep.libraries:
            data["library_id"].append(library.id)
            data["protocol_id"].append(library.protocol_id)

        df = pd.DataFrame(data)

        if df["protocol_id"].isna().any() and self.lab_prep.prep_file is not None:
            protocols = db.pd.get_protocol_kits()
            protocols = (
                protocols.groupby(['protocol_id', 'combination_num'])
                .agg(identifiers=('kit_identifier', lambda x: ';'.join(sorted(x))))
                .reset_index()
            )
            
            path = os.path.join(runtime.app.media_folder, self.lab_prep.prep_file.path)
            wb = openpyxl.load_workbook(path)
            active_sheet = wb["prep_table"]
            
            column_mapping: dict[str, str] = {}
            for col_i in range(1, min(active_sheet.max_column, 96)):
                col = get_column_letter(col_i + 1)
                column_name = active_sheet[f"{col}1"].value
                column_mapping[column_name] = col

            for row_i in range(2, active_sheet.max_row + 1):
                library_id = active_sheet[f"{column_mapping['library_id']}{row_i}"].value
                library_kits = active_sheet[f"{column_mapping['library_kits']}{row_i}"].value

                protocol_id = None
                if pd.notna(library_kits) and (library_kits := str(library_kits).strip()):
                    library_kits = ";".join(sorted([kit.strip().removeprefix("#").strip() for kit in library_kits.split(";")]))
                    if library_kits in protocols["identifiers"].values:
                        protocol_id = protocols.loc[protocols["identifiers"] == library_kits, "protocol_id"].iloc[0]
                
                df.loc[df["library_id"] == library_id, "protocol_id"] = protocol_id

        return df

    def validate(self) -> bool:
        if not super().validate():
            return False
        
        if not self.spreadsheet.validate():
            return False
        
        self.df = self.spreadsheet.df

        if len(self.spreadsheet._errors) > 0:
            return False

        return True

    def process_request(self) -> Response:
        if not self.validate():
            return self.make_response()

        for _, row in self.df.iterrows():
            library = db.libraries[int(row["library_id"])]
            library.protocol_id = int(row["protocol_id"]) if pd.notna(row["protocol_id"]) else None
            db.libraries.update(library)
        
        flash("Protocols Submitted!", "success")
        return make_response(redirect=(url_for("lab_preps_page.lab_prep", lab_prep_id=self.lab_prep.id, tab="lab_prep-checklist-tab")))
        

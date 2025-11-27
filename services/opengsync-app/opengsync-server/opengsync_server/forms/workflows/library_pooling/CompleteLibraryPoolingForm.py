import os

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from flask import Response, url_for, flash
from flask_htmx import make_response

from opengsync_db import models
from opengsync_db.categories import PoolType, SeqRequestStatus, LibraryStatus

from .... import logger, db
from ....tools import utils
from ....core.RunTime import runtime
from ...MultiStepForm import MultiStepForm


class CompleteLibraryPoolingForm(MultiStepForm):
    _template_path = "workflows/library_pooling/complete-pooling.html"
    _workflow_name = "library_pooling"
    _step_name = "complete_library_pooling"

    def __init__(self, lab_prep: models.LabPrep, uuid: str | None, formdata: dict | None = None):
        MultiStepForm.__init__(
            self, workflow=CompleteLibraryPoolingForm._workflow_name,
            step_name=CompleteLibraryPoolingForm._step_name, uuid=uuid,
            formdata=formdata, step_args={"lab_prep": lab_prep},
        )
        self.lab_prep = lab_prep
        self._context["lab_prep"] = lab_prep
        self.pooling_table = self.tables["pooling_table"]
        self.barcode_table = db.pd.get_lab_prep_barcodes(self.lab_prep.id)
        self.barcode_table["pool"] = utils.map_columns(self.barcode_table, self.pooling_table, "library_id", "pool")
        self.barcode_table = utils.check_indices(self.barcode_table, groupby="pool")

    def validate(self) -> bool:
        validated = super().validate()
        return validated

    def process_request(self, user: models.User) -> Response:
        if not self.validate():
            self.prepare()
            return self.make_response()
        
        library_table = self.tables["library_table"]
        self.pooling_table["old_pool_id"] = utils.map_columns(self.pooling_table, library_table, "library_id", "pool_id")
        self.pooling_table["experiment_id"] = None

        for pool in self.lab_prep.pools:
            self.pooling_table.loc[self.pooling_table["old_pool_id"] == pool.id, "experiment_id"] = pool.experiment_id
            db.pools.delete(pool.id)

        if len(self.pooling_table["pool"].unique()) == 1:
            self.pooling_table["pool"] = "1"
        
        # if all the experiment_ids are the same in the pool we can link it with the experiment
        experiment_mappings = {}
        for pool_suffix, _df in self.pooling_table.groupby("pool"):
            if len(_df["experiment_id"].unique()) == 1 and pd.notna(_df["experiment_id"].iloc[0]):
                experiment_mappings[pool_suffix] = _df["experiment_id"].iloc[0]
                    
        if len(self.pooling_table["pool"].unique()) > 1:
            for pool_suffix, df in self.pooling_table.groupby("pool"):
                if pool_suffix == "t" or pool_suffix == "skip":
                    continue
                if pool_suffix == "x":
                    for _, row in df.iterrows():
                        library = db.libraries[int(row["library_id"])]
                        library.status = LibraryStatus.FAILED
                        db.libraries.update(library)
                    continue

                pool_suffix = str(pool_suffix).removeprefix(f"{self.lab_prep.name}_").strip()
                pool = db.pools.create(
                    name=f"{self.lab_prep.name}_{pool_suffix}", pool_type=PoolType.INTERNAL,
                    contact_email=user.email, contact_name=user.name, owner_id=user.id,
                    lab_prep_id=self.lab_prep.id, experiment_id=experiment_mappings.get(pool_suffix, None),
                )
                for _, row in df.iterrows():
                    library = db.libraries[int(row["library_id"])]
                    library.pool_id = pool.id
                    library.status = LibraryStatus.POOLED
                    db.libraries.update(library)
                    
        elif len(self.pooling_table["pool"].unique()) > 0:
            pool = db.pools.create(
                name=self.lab_prep.name, pool_type=PoolType.INTERNAL,
                contact_email=user.email, contact_name=user.name, owner_id=user.id,
                lab_prep_id=self.lab_prep.id, experiment_id=experiment_mappings.get("1", None)
            )
            for _, row in self.pooling_table.iterrows():
                library = db.libraries[int(row["library_id"])]
                library.pool_id = pool.id
                library.status = LibraryStatus.POOLED
                db.libraries.update(library)

        request_ids = set()

        if self.lab_prep.prep_file is not None:
            path = os.path.join(runtime.app.media_folder, self.lab_prep.prep_file.path)
            wb = openpyxl.load_workbook(path)
            active_sheet = wb["prep_table"]
            
            column_mapping: dict[str, str] = {}
            for col_i in range(1, min(active_sheet.max_column, 96)):
                col = get_column_letter(col_i + 1)
                column_name = active_sheet[f"{col}1"].value
                column_mapping[column_name] = col

            for i, ((library_id, pool), _) in enumerate(self.pooling_table.groupby(["library_id", "pool"], dropna=False)):
                library = db.libraries[int(library_id)]
                request_ids.add(library.seq_request_id)
                active_sheet[f"{column_mapping['pool']}{i + 2}"].value = pool

            wb.save(path)

        for request_id in request_ids:
            if (seq_request := db.seq_requests.get(request_id)) is None:
                logger.error(f"{self.uuid}: SeqRequest {request_id} not found")
                raise ValueError(f"{self.uuid}: SeqRequest {request_id} not found")
            
            prepared = True
            for library in seq_request.libraries:
                prepared = prepared and library.status.id >= LibraryStatus.POOLED.id

            if prepared and seq_request.status == SeqRequestStatus.ACCEPTED:
                seq_request.status = SeqRequestStatus.PREPARED
                db.seq_requests.update(seq_request)

        self.complete()
        flash("Library Indexing completed!", "success")
        return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=self.lab_prep.id))

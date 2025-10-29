import os

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from flask import Response, url_for, flash
from flask_htmx import make_response

from opengsync_db import models
from opengsync_db.categories import IndexType, BarcodeOrientation

from .... import logger, tools, db  # noqa F401
from ....core import exceptions, runtime
from ...MultiStepForm import MultiStepForm


class CompleteReindexForm(MultiStepForm):
    _template_path = "workflows/reindex/reindex-complete.html"
    _workflow_name = "reindex"
    _step_name = "complete_reindex"

    def __init__(
        self,
        seq_request: models.SeqRequest | None,
        lab_prep: models.LabPrep | None,
        pool: models.Pool | None,
        formdata: dict | None,
        uuid: str | None
    ):
        MultiStepForm.__init__(
            self, workflow=CompleteReindexForm._workflow_name,
            formdata=formdata, step_name=CompleteReindexForm._step_name,
            uuid=uuid, step_args={"seq_request": seq_request, "lab_prep": lab_prep, "pool": pool},
        )
        self.seq_request = seq_request
        self.lab_prep = lab_prep
        self.pool = pool
        self.library_table = self.tables["library_table"]
        self.barcode_table = self.tables["barcode_table"]
        self.barcode_table["orientation_id"] = self.barcode_table["orientation_i7_id"]
        self.barcode_table = tools.check_indices(self.barcode_table)

        if (tenx_atac_barcode_table := self.tables.get("tenx_atac_barcode_table")) is not None:
            self.barcode_table = pd.concat([self.barcode_table, tenx_atac_barcode_table], ignore_index=True)
        
        self.barcode_table.loc[
            pd.notna(self.barcode_table["orientation_i7_id"]) &
            (self.barcode_table["orientation_i7_id"] != self.barcode_table["orientation_i5_id"]),
            "orientation_id"
        ] = None
        
        self.index_col = self.metadata["index_col"]
        self._context["barcode_table"] = self.barcode_table

        self.url_context = {}
        if seq_request is not None:
            self._context["seq_request"] = seq_request
            self.url_context["seq_request_id"] = seq_request.id
        if lab_prep is not None:
            self._context["lab_prep"] = lab_prep
            self.url_context["lab_prep_id"] = lab_prep.id
        if pool is not None:
            self._context["pool"] = pool
            self.url_context["pool_id"] = pool.id

        self.post_url = url_for("reindex_workflow.complete_reindex", uuid=self.uuid, **self.url_context)

    def process_request(self) -> Response:
        if not self.validate():
            return self.make_response()
        
        barcode_table = self.tables["barcode_table"]
        tenx_atac_barcode_table = self.tables.get("tenx_atac_barcode_table")
        
        for (library_id, index_type_id), _ in self.barcode_table.groupby(["library_id", "index_type_id"], dropna=False):
            library = db.libraries[int(library_id)]

            try:
                index_type_id = int(index_type_id)
                index_type = IndexType.get(index_type_id)
            except ValueError:
                logger.error(f"{self.uuid}: Invalid index_type_id {index_type_id} for library {library_id}")
                raise exceptions.InternalServerErrorException(f"{self.uuid}: Invalid index_type_id {index_type_id} for library {library_id}")

            library = db.libraries.remove_indices(library_id=library.id)
            
            library.index_type = index_type
            db.libraries.update(library)

            match index_type:
                case IndexType.TENX_ATAC_INDEX:
                    if tenx_atac_barcode_table is None:
                        logger.error(f"{self.uuid}: TENX_ATAC_INDEX selected but no tenx_atac_barcode_table found.")
                        raise exceptions.InternalServerErrorException(f"{self.uuid}: TENX_ATAC_INDEX selected but no tenx_atac_barcode_table found.")
                    
                    df = tenx_atac_barcode_table[tenx_atac_barcode_table["library_id"] == library.id]
                    if index_type == IndexType.TENX_ATAC_INDEX:
                        if len(df) != 4:
                            logger.warning(f"{self.uuid}: Expected 4 barcodes (i7) for index type {library.index_type}, found {len(df)}.")
                case _:
                    df = barcode_table[barcode_table["library_id"] == library.id]
                    if len(df) != 1:
                        logger.warning(f"{self.uuid}: Expected 1 barcode for index type {library.index_type}, found {len(df)}.")

            for _, row in df.iterrows():
                if index_type == IndexType.TENX_ATAC_INDEX:
                    for i in range(1, 5):
                        if pd.isna(row[f"sequence_{i}"]):
                            logger.error(f"{self.uuid}: Missing sequence_{i} for TENX_ATAC_INDEX in library {row['library_name']}.")
                            raise ValueError(f"Missing sequence_{i} for TENX_ATAC_INDEX in library {row['library_name']}.")
                        
                        library = db.libraries.add_index(
                            library_id=library.id,
                            index_kit_i7_id=int(row["kit_id"]) if pd.notna(row["kit_id"]) else None,
                            index_kit_i5_id=None,
                            name_i7=row["name"] or None,
                            name_i5=None,
                            sequence_i7=row[f"sequence_{i}"],
                            sequence_i5=None,
                            orientation=BarcodeOrientation.FORWARD if pd.notna(row["kit_id"]) else None,
                        )
                else:
                    orientation = None
                    if pd.notna(row["orientation_i7_id"]):
                        orientation = BarcodeOrientation.get(row["orientation_i7_id"])

                    if orientation is not None and pd.notna(row["orientation_i5_id"]):
                        if orientation.id != row["orientation_i5_id"]:
                            logger.error(f"{self.uuid}: Conflicting orientations for i7 and i5 in library {row['library_name']}.")
                            raise ValueError("Conflicting orientations for i7 and i5.")
                    library = db.libraries.add_index(
                        library_id=library.id,
                        index_kit_i7_id=int(row["kit_i7_id"]) if pd.notna(row["kit_i7_id"]) else None,
                        index_kit_i5_id=int(row["kit_i5_id"]) if pd.notna(row["kit_i5_id"]) else None,
                        name_i7=row["name_i7"] if pd.notna(row["name_i7"]) else None,
                        name_i5=row["name_i5"] if pd.notna(row["name_i5"]) else None,
                        sequence_i7=row["sequence_i7"],
                        sequence_i5=row["sequence_i5"] if pd.notna(row["sequence_i5"]) else None,
                        orientation=orientation,
                    )

        if self.lab_prep is not None and self.lab_prep.prep_file is not None:
            path = os.path.join(runtime.app.media_folder, self.lab_prep.prep_file.path)
            wb = openpyxl.load_workbook(path)
            active_sheet = wb["prep_table"]
            
            column_mapping: dict[str, str] = {}
            for col_i in range(1, min(active_sheet.max_column, 96)):
                col = get_column_letter(col_i + 1)
                column_name = active_sheet[f"{col}1"].value
                column_mapping[column_name] = col

            for i, ((library_id, index_type_id), _) in enumerate(self.barcode_table.groupby(["library_id", "index_type_id"], dropna=False)):
                library = db.libraries[int(library_id)]
                
                try:
                    index_type_id = int(index_type_id)
                    index_type = IndexType.get(index_type_id)
                except ValueError:
                    logger.error(f"{self.uuid}: Invalid index_type_id {index_type_id} for library {library_id}")
                    raise exceptions.InternalServerErrorException(f"{self.uuid}: Invalid index_type_id {index_type_id} for library {library_id}")

                if index_type == IndexType.TENX_ATAC_INDEX:
                    if tenx_atac_barcode_table is None:
                        logger.error(f"{self.uuid}: TENX_ATAC_INDEX selected but no tenx_atac_barcode_table found.")
                        raise exceptions.InternalServerErrorException(f"{self.uuid}: TENX_ATAC_INDEX selected but no tenx_atac_barcode_table found.")
                    
                    active_sheet[f"{column_mapping['sequence_i7']}{i + 2}"].value = library.sequences_i7_str(sep=";")
                    active_sheet[f"{column_mapping['sequence_i5']}{i + 2}"].value = None
                    active_sheet[f"{column_mapping['name_i7']}{i + 2}"].value = ";".join(set([index.name_i7 for index in library.indices if index.name_i7]))
                    active_sheet[f"{column_mapping['name_i5']}{i + 2}"].value = None
                    active_sheet[f"{column_mapping['pool']}{i + 2}"].value = library.pool.name if library.pool is not None else None
                    active_sheet[f"{column_mapping['kit_i7']}{i + 2}"].value = ";".join(set([index.index_kit_i7.identifier for index in library.indices if index.index_kit_i7]))
                    active_sheet[f"{column_mapping['kit_i5']}{i + 2}"].value = None
                    active_sheet[f"{column_mapping['index_well']}{i + 2}"].value = next(iter(barcode_table[barcode_table["library_id"] == library.id]["index_well"].tolist()), None)
                else:

                    active_sheet[f"{column_mapping['sequence_i7']}{i + 2}"].value = library.sequences_i7_str(sep=";")
                    active_sheet[f"{column_mapping['sequence_i5']}{i + 2}"].value = library.sequences_i5_str(sep=";")
                    active_sheet[f"{column_mapping['name_i7']}{i + 2}"].value = ";".join([index.name_i7 for index in library.indices if index.name_i7])
                    active_sheet[f"{column_mapping['name_i5']}{i + 2}"].value = ";".join([index.name_i5 for index in library.indices if index.name_i5])
                    active_sheet[f"{column_mapping['pool']}{i + 2}"].value = library.pool.name if library.pool is not None else None
                    active_sheet[f"{column_mapping['kit_i7']}{i + 2}"].value = ";".join([index.index_kit_i7.identifier for index in library.indices if index.index_kit_i7])
                    active_sheet[f"{column_mapping['kit_i5']}{i + 2}"].value = ";".join([index.index_kit_i5.identifier for index in library.indices if index.index_kit_i5])
                    active_sheet[f"{column_mapping['index_well']}{i + 2}"].value = next(iter(barcode_table[barcode_table["library_id"] == library.id]["index_well"].tolist()), None)

            logger.debug(f"Overwriting existing file: {os.path.join(runtime.app.media_folder, self.lab_prep.prep_file.path)}")
            wb.save(path)

        flash("Libraries Re-Indexed!", "success")
        self.complete()
        
        if self.seq_request is not None:
            return make_response(redirect=url_for("seq_requests_page.seq_request", seq_request_id=self.seq_request.id))
        if self.lab_prep is not None:
            return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=self.lab_prep.id))
        if self.pool is not None:
            return make_response(redirect=url_for("pools_page.pool", pool_id=self.pool.id))
        
        return make_response(redirect=url_for("dashboard"))
        
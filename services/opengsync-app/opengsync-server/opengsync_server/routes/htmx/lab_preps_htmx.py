import os
import io
import json
from typing import Literal

import pandas as pd

import openpyxl
from openpyxl import styles as openpyxl_styles
from openpyxl.utils import get_column_letter

from flask import Blueprint, render_template, request, flash, url_for, Response
from flask_htmx import make_response

from opengsync_db import models, PAGE_LIMIT
from opengsync_db.categories import LabChecklistType, PoolStatus, LibraryStatus, PrepStatus, SeqRequestStatus, LibraryType, SampleStatus

from ... import db, forms, logger
from ...core import wrappers, exceptions
from ...core.RunTime import runtime
from ...tools.spread_sheet_components import TextColumn
from ...tools import StaticSpreadSheet


lab_preps_htmx = Blueprint("lab_preps_htmx", __name__, url_prefix="/htmx/lab_preps/")


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get(current_user: models.User, page: int = 0):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    sort_by = request.args.get("sort_by", "id")
    sort_order = request.args.get("sort_order", "desc")
    descending = sort_order == "desc"
    offset = PAGE_LIMIT * page

    if (status_in := request.args.get("status_id_in")) is not None:
        status_in = json.loads(status_in)
        try:
            status_in = [PrepStatus.get(int(status)) for status in status_in]
        except ValueError:
            raise exceptions.BadRequestException()
    
        if len(status_in) == 0:
            status_in = None

    if (protocol_in := request.args.get("protocol_id_in")) is not None:
        protocol_in = json.loads(protocol_in)
        try:
            protocol_in = [LabChecklistType.get(int(protocol_)) for protocol_ in protocol_in]
        except ValueError:
            raise exceptions.BadRequestException()
    
        if len(protocol_in) == 0:
            protocol_in = None

    lab_preps, n_pages = db.lab_preps.find(
        status_in=status_in, checklist_type_in=protocol_in,
        offset=offset, limit=PAGE_LIMIT, sort_by=sort_by, descending=descending, count_pages=True
    )
    
    return render_template(
        "components/tables/lab_prep.html", lab_preps=lab_preps, n_pages=n_pages, active_page=page,
        sort_by=sort_by, sort_order=sort_order, status_in=status_in, protocol_in=protocol_in,
    )


@wrappers.htmx_route(lab_preps_htmx, db=db)
def table_query():
    if (word := request.args.get("name")) is not None:
        field_name = "name"
    elif (word := request.args.get("id")) is not None:
        field_name = "id"
    elif (word := request.args.get("creator_id")) is not None:
        field_name = "creator_id"
    else:
        raise exceptions.BadRequestException()

    if (status_in := request.args.get("status_id_in")) is not None:
        status_in = json.loads(status_in)
        try:
            status_in = [PrepStatus.get(int(status)) for status in status_in]
        except ValueError:
            raise exceptions.BadRequestException()
    
        if len(status_in) == 0:
            status_in = None

    if (checklist_type_in := request.args.get("checklist_type_id_in")) is not None:
        checklist_type_in = json.loads(checklist_type_in)
        try:
            checklist_type_in = [LabChecklistType.get(int(checklist_type)) for checklist_type in checklist_type_in]
        except ValueError:
            raise exceptions.BadRequestException()
    
        if len(checklist_type_in) == 0:
            checklist_type_in = None

    lab_preps: list[models.LabPrep] = []
    if field_name == "name":
        lab_preps = db.lab_preps.query(name=word, checklist_type_in=checklist_type_in, status_in=status_in)
    elif field_name == "id":
        try:
            _id = int(word)
            if (lab_prep := db.lab_preps.get(_id)) is not None:
                lab_preps.append(lab_prep)
        except ValueError:
            pass
    elif field_name == "creator_id":
        lab_preps = db.lab_preps.query(creator=word, checklist_type_in=checklist_type_in, status_in=status_in)

    return make_response(
        render_template(
            "components/tables/lab_prep.html",
            current_query=word, active_query_field=field_name,
            lab_preps=lab_preps, checklist_type_in=checklist_type_in, status_in=status_in
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get_form(current_user: models.User, form_type: Literal["create", "edit"]):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if form_type not in ["create", "edit"]:
        raise exceptions.BadRequestException()
    
    if (lab_prep_id := request.args.get("lab_prep_id")) is not None:
        try:
            lab_prep_id = int(lab_prep_id)
        except ValueError:
            raise exceptions.BadRequestException()
        
        if form_type != "edit":
            raise exceptions.BadRequestException()
        
        if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
            raise exceptions.NotFoundException()
        
        return forms.models.LabPrepForm(form_type=form_type, lab_prep=lab_prep).make_response()
    
    # seq_request_id must be provided if form_type is "edit"
    if form_type == "edit":
        raise exceptions.BadRequestException()

    return forms.models.LabPrepForm(form_type=form_type).make_response()


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def create(current_user: models.User):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    form = forms.models.LabPrepForm(formdata=request.form, form_type="create")
    return form.process_request(current_user)


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def edit(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    form = forms.models.LabPrepForm(formdata=request.form, form_type="edit", lab_prep=lab_prep)
    return form.process_request(current_user)


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def complete(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    for pool in lab_prep.pools:
        if pool.status < PoolStatus.STORED:
            pool.status = PoolStatus.STORED
            db.pools.update(pool)

    for library in lab_prep.libraries:
        is_prepared = True
        for sr_library in library.seq_request.libraries:
            is_prepared = sr_library.status >= LibraryStatus.POOLED and is_prepared
            if not is_prepared:
                break
        if is_prepared:
            library.seq_request.status = SeqRequestStatus.PREPARED
            db.libraries.update(library)

    lab_prep.status = PrepStatus.COMPLETED
    db.lab_preps.update(lab_prep)

    flash("Lab prep completed!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["DELETE"])
def delete(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if lab_prep.status != PrepStatus.PREPARING:
        flash("Cannot delete completed prep.", "warning")
        return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))
    
    db.lab_preps.delete(lab_prep_id=lab_prep.id)
    flash("Lab prep deleted!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_preps"))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def uncomplete(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()

    lab_prep.status = PrepStatus.PREPARING
    db.lab_preps.update(lab_prep)

    flash("Lab prep completed!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["DELETE"], arg_params=["library_id"])
def remove_library(current_user: models.User, lab_prep_id: int, library_id: int, page: int = 0):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if lab_prep.status != PrepStatus.PREPARING:
        raise exceptions.BadRequestException()
    
    db.lab_preps.remove_library(lab_prep_id=lab_prep.id, library_id=library_id)

    sort_by = request.args.get("sort_by", "id")
    sort_order = request.args.get("sort_order", "desc")
    descending = sort_order == "desc"

    libraries, n_pages = db.libraries.find(page=page, limit=PAGE_LIMIT, lab_prep_id=lab_prep_id, sort_by=sort_by, descending=descending)
    return make_response(
        render_template(
            "components/tables/lab_prep-library.html",
            libraries=libraries, n_pages=n_pages, active_page=page,
            sort_by=sort_by, sort_order=sort_order, lab_prep=lab_prep
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET"])
def download_template(current_user: models.User, lab_prep_id: int, direction: Literal["rows", "columns"]) -> Response:
    if direction not in ("rows", "columns"):
        raise exceptions.BadRequestException()
    
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()

    if runtime.app.static_folder is None:
        logger.error("Static folder not set")
        raise ValueError("Static folder not set")

    filepath = os.path.join(runtime.app.static_folder, "resources", "templates", "library_prep", lab_prep.checklist_type.prep_file_name)

    if not os.path.exists(filepath):
        logger.error(f"File not found: {filepath}")
        raise exceptions.NotFoundException()

    template = openpyxl.load_workbook(filepath)

    n = 12 if direction == "rows" else 8
    pattern = [True] * n + [False] * n

    def if_color(i):
        return pattern[i]

    prep_table = template["prep_table"]
    column_mapping: dict[str, str] = {}
    
    for col_i in range(0, min(prep_table.max_column, 96)):
        col = get_column_letter(col_i + 1)
        column_name = prep_table[f"{col}1"].value
        column_mapping[column_name] = col
        
        for row_idx, cell in enumerate(prep_table[col][1:]):
            if if_color(row_idx % (n * 2)):
                cell.fill = openpyxl_styles.PatternFill(start_color="ced4da", end_color="ced4da", fill_type="solid")
            else:
                cell.fill = openpyxl_styles.PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

    for row_idx, cell in enumerate(prep_table[column_mapping["plate_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")

    for row_idx, cell in enumerate(prep_table[column_mapping["index_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")
        
    for i, library in enumerate(lab_prep.libraries):
        library_id_cell = prep_table[f"{column_mapping['library_id']}{i + 2}"]
        library_name_cell = prep_table[f"{column_mapping['library_name']}{i + 2}"]
        requestor_cell = prep_table[f"{column_mapping['requestor']}{i + 2}"]
        sequence_i7_cell = prep_table[f"{column_mapping['sequence_i7']}{i + 2}"]
        sequence_i5_cell = prep_table[f"{column_mapping['sequence_i5']}{i + 2}"]
        # kit_i7_cell = active_sheet[f"{column_mapping['kit_i7']}{i + 2}"]
        # kit_i5_cell = active_sheet[f"{column_mapping['kit_i5']}{i + 2}"]
        name_i7_cell = prep_table[f"{column_mapping['name_i7']}{i + 2}"]
        name_i5_cell = prep_table[f"{column_mapping['name_i5']}{i + 2}"]
        library_id_cell.value = library.id
        library_name_cell.value = library.name
        requestor_cell.value = library.seq_request.requestor.name
        if len(library.indices) > 0:
            name_i7_cell.value = library.indices[0].name_i7
            name_i5_cell.value = library.indices[0].name_i5
        sequence_i7_cell.value = library.sequences_i7_str(";")
        sequence_i5_cell.value = library.sequences_i5_str(";")

    if "FLEX_table" in template.sheetnames:
        flex_table = template["FLEX_table"]
        column_mapping: dict[str, str] = {}
    
        for col_i in range(0, min(flex_table.max_column, 96)):
            col = get_column_letter(col_i + 1)
            column_name = flex_table[f"{col}1"].value
            column_mapping[column_name] = col

        i = 2
        for library in lab_prep.libraries:
            if library.type == LibraryType.TENX_SC_GEX_FLEX:
                for sample_link in library.sample_links:
                    sample_num_cell = flex_table[f"{column_mapping['sample_num']}{i}"]
                    sample_name_cell = flex_table[f"{column_mapping['sample_name']}{i}"]
                    sample_num_cell.value = i - 1
                    sample_name_cell.value = sample_link.sample.name
                    i += 1

    if "10X_table" in template.sheetnames:
        tenx_table = template["10X_table"]
        column_mapping: dict[str, str] = {}
    
        for col_i in range(0, min(tenx_table.max_column, 96)):
            col = get_column_letter(col_i + 1)
            column_name = tenx_table[f"{col}1"].value
            column_mapping[column_name] = col

        i = 2
        for library in lab_prep.libraries:
            if library.type in (LibraryType.TENX_SC_GEX_3PRIME, LibraryType.TENX_SC_GEX_5PRIME):
                for sample_link in library.sample_links:
                    sample_num_cell = tenx_table[f"{column_mapping['sample_num']}{i}"]
                    sample_name_cell = tenx_table[f"{column_mapping['sample_name']}{i}"]
                    library_name_cell = tenx_table[f"{column_mapping['library_name']}{i}"]
                    library_id_cell = tenx_table[f"{column_mapping['library_id']}{i}"]
                    requestor_cell = tenx_table[f"{column_mapping['requestor']}{i}"]

                    sample_num_cell.value = i - 1
                    sample_name_cell.value = sample_link.sample.name
                    library_name_cell.value = library.name
                    library_id_cell.value = library.id
                    requestor_cell.value = library.seq_request.requestor.name
                    i += 1
            
    bytes_io = io.BytesIO()
    template.save(bytes_io)

    bytes_io.seek(0)
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
    return Response(
        bytes_io, mimetype=mimetype,
        headers={"Content-disposition": f"attachment; filename={lab_prep.name}_{lab_prep.checklist_type.abbreviation}_{direction}.xlsx"}
    )


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def prep_table_upload_form(current_user: models.User, lab_prep_id: int) -> Response:
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if request.method == "GET":
        form = forms.workflows.library_prep.LibraryPrepForm(lab_prep=lab_prep)
        return form.make_response()
    else:
        form = forms.workflows.library_prep.LibraryPrepForm(lab_prep=lab_prep, formdata=request.form | request.files)
        return form.process_request(user=current_user)
    

@wrappers.htmx_route(lab_preps_htmx, db=db)
def checklist(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    checklist = lab_prep.get_checklist()
    return make_response(
        render_template(
            "components/checklists/lab_prep.html",
            lab_prep=lab_prep, **checklist
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get_files(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()

    return make_response(
        render_template(
            "components/file-list.html",
            files=lab_prep.media_files, delete="lab_preps_htmx.delete_file",
            delete_context={"lab_prep_id": lab_prep_id}
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["DELETE"])
def delete_file(current_user: models.User, lab_prep_id: int, file_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if (file := db.media_files.get(file_id)) is None:
        raise exceptions.NotFoundException()
    
    if file not in lab_prep.media_files:
        raise exceptions.BadRequestException()

    file_path = os.path.join(runtime.app.media_folder, file.path)
    if os.path.exists(file_path):
        os.remove(file_path)
    db.media_files.delete(file_id=file.id)

    logger.info(f"Deleted file '{file.name}' from prep (id='{lab_prep_id}')")
    flash(f"Deleted file '{file.name}' from prep.", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def file_form(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if request.method == "GET":
        form = forms.file.LabPrepAttachmentForm(lab_prep=lab_prep)
        return form.make_response()
    elif request.method == "POST":
        form = forms.file.LabPrepAttachmentForm(lab_prep=lab_prep, formdata=request.form | request.files)
        return form.process_request(current_user)
    else:
        raise exceptions.MethodNotAllowedException()
    

@wrappers.htmx_route(lab_preps_htmx, db=db)
def plates_tab(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    return make_response(render_template("components/plates.html", plates=lab_prep.plates, from_page=f"lab_prep@{lab_prep.id}"))
    

@wrappers.htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def comment_form(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if request.method == "GET":
        form = forms.comment.LabPrepCommentForm(lab_prep=lab_prep)
        return form.make_response()
    elif request.method == "POST":
        form = forms.comment.LabPrepCommentForm(lab_prep=lab_prep, formdata=request.form)
        return form.process_request(current_user)
    else:
        raise exceptions.MethodNotAllowedException()


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get_comments(current_user: models.User, lab_prep_id: int):
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()

    return make_response(
        render_template(
            "components/comment-list.html",
            comments=lab_prep.comments
        )
    )


@wrappers.htmx_route(lab_preps_htmx, db=db)
def get_mux_table(current_user: models.User, lab_prep_id: int):
    if not current_user.is_insider():
        raise exceptions.NoPermissionsException()
    
    if (lab_prep := db.lab_preps.get(lab_prep_id)) is None:
        raise exceptions.NotFoundException()
    
    df = db.pd.get_lab_prep_pooling_table(lab_prep.id)

    df = df.sort_values(by=["library_name", "sample_pool", "sample_name"])

    mux_table = {
        "sample_name": [],
        "library_name": [],
        "sample_pool": [],
        "barcode": [],
        "pattern": [],
        "read": []
    }

    for _, row in df.iterrows():
        if row["mux_type_id"] is None:
            continue
        
        mux_table["sample_name"].append(row["sample_name"])
        mux_table["library_name"].append(row["library_name"])
        mux_table["sample_pool"].append(row["sample_pool"])
        if (mux := row.get("mux")) is None:
            mux_table["barcode"].append(None)
            mux_table["pattern"].append(None)
            mux_table["read"].append(None)
        else:
            mux_table["barcode"].append(mux.get("barcode"))
            mux_table["pattern"].append(mux.get("pattern"))
            mux_table["read"].append(mux.get("read"))

    df = pd.DataFrame(mux_table)
    if df["pattern"].isna().all():
        df = df.drop(columns=["pattern"])
    if df["read"].isna().all():
        df = df.drop(columns=["read"])
        
    columns = []
    for col in df.columns:
        columns.append(
            TextColumn(
                col,
                col.replace("_", " ").title().replace("Id", "ID").replace("Cmo", "CMO"),
                {
                    "sample_name": 250,
                    "library_name": 300,
                    "sample_pool": 200,
                    "barcode": 100,
                    "read": 80,
                    "pattern": 150
                }.get(col, 100),
                max_length=1000
            )
        )

    spreadsheet = StaticSpreadSheet(df, columns=columns, id=f"lab_prep_mux_table-{lab_prep_id}")

    return make_response(render_template("components/itable.html", spreadsheet=spreadsheet))


import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment, Side, Border, Color
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell, MergedCell
from openpyxl import Workbook
from colorsys import rgb_to_hls, hls_to_rgb

RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:


# https://stackoverflow.com/questions/58429823/getting-excel-cell-background-themed-color-as-hex-with-openpyxl
def get_theme_colors(wb: Workbook):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]

    colors = []

    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        for i in list(accent): # walk all child nodes, rather than assuming [0]
            if 'window' in i.attrib['val']:
                colors.append(i.attrib['lastClr'])
            else:
                colors.append(i.attrib['val'])

    return colors

def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()

def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))
    
def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

def theme_and_tint_to_rgb(themes, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = themes[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))


def argb_hex_to_rgb_hex(argb_str):
    # Remove alpha (first 2 chars)
    rgb_str = argb_str[2:]
    # Convert to integers
    r = int(rgb_str[0:2], 16)
    g = int(rgb_str[2:4], 16)
    b = int(rgb_str[4:6], 16)
    return f"{r:02X}{g:02X}{b:02X}"

def extract_style(theme_colors: list[str], cell: Cell | MergedCell) -> dict:
    """Extract cell style and convert to Univer format"""
    style = {}
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    
    # Font family (ff)
    if font and font.name:
        style["ff"] = font.name
    
    # Font size (fs)
    if font and font.size:
        style["fs"] = int(font.size)
    
    # Italic (it)
    if font and font.italic:
        style["it"] = 1
    
    # Bold (bl)
    if font and font.b:
        style["bl"] = 1
    
    # Underline (ul)
    if font and font.underline and font.underline != "none":
        style["ul"] = 1
    
    # Strikethrough (st)
    if font and font.strike:
        style["st"] = 1

    # Font color (cl)
    if font and font.color:
        font.color: Color  # type: ignore
        if font.color.type == "rgb":
            style["cl"] = { "rgb": f"#{argb_hex_to_rgb_hex(font.color.rgb)}" }
        elif font.color.type == "theme":
            style["cl"] = { "rgb": f"#{theme_and_tint_to_rgb(theme_colors, font.color.theme, font.color.tint)}" }

    # Background color (bg)
    if fill and fill.fill_type and fill.fill_type != "none":
        fill: PatternFill
        if fill.fgColor.type == "rgb":
            style["bg"] = { "rgb": f"#{argb_hex_to_rgb_hex(fill.fgColor.rgb)}" }
        elif fill.fgColor.type == "theme":
            style["bg"] = { "rgb": f"#{theme_and_tint_to_rgb(theme_colors, fill.fgColor.theme, fill.fgColor.tint)}" }
    
    # Superscript/subscript (va)
    # openpyxl doesn't directly support this, so skip
    
    # Text rotation (tr)
    if alignment and alignment.text_rotation:
        style["tr"] = int(alignment.text_rotation)
    
    # Horizontal alignment (ht)
    if alignment and alignment.horizontal:
        # Map openpyxl values to Univer format
        ht_map = {
            "left": "left",
            "center": "center", 
            "right": "right",
            "justify": "justify",
            "fill": "fill",
            "distributed": "distributed"
        }
        if alignment.horizontal in ht_map:
            style["ht"] = ht_map[alignment.horizontal]
    
    # Vertical alignment (vt)
    if alignment and alignment.vertical:
        # Map openpyxl values to Univer format
        vt_map = {
            "top": "top",
            "center": "middle",
            "bottom": "bottom",
            "justify": "justify",
            "distributed": "distributed"
        }
        if alignment.vertical in vt_map:
            style["vt"] = vt_map[alignment.vertical]
    
    # Truncate overflow (tb) - not directly supported in openpyxl
    # This relates to text wrapping, so we can infer from wrap_text
    if alignment and alignment.wrap_text is False:
        style["tb"] = 1  # truncate if not wrapping
    
    # Padding (pd) - not directly supported in openpyxl, skip
    
    # Number format (n)
    if cell.number_format and cell.number_format != "General":
        style["n"] = cell.number_format
    
    return style


def xlsx_to_univer_snapshot(path: str) -> tuple[dict, dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    snapshot = {
        "id": "workbook",
        "name": "Workbook",
        "sheetOrder": wb.sheetnames,
        "sheets": {},
    }
    col_style = {}

    theme_colors: list[str] = get_theme_colors(wb)

    for col_idx, sheet in enumerate(wb.sheetnames):
        ws: Worksheet = wb[sheet]
        snapshot["sheets"][sheet] = {
            "id": sheet,
            "name": sheet,
            "columnCount": len(list(ws.columns)),
            "cellData": {}, # (row, col)
        }
        sheet_col_style = {}
        for col_idx, letter in enumerate(list(ws.column_dimensions)):
            sheet_col_style[col_idx] = { "width": ws.column_dimensions[letter].width * 7.5 if ws.column_dimensions[letter].width else 64 }
        
        col_style[sheet] = sheet_col_style

        for row_idx, row in enumerate(ws.iter_rows()):
            row_data = {}
            for cell_idx, cell in enumerate(row):
                row_data[cell_idx] = {
                    "v": str(cell.value) if cell.value is not None else "",
                    "s": extract_style(theme_colors, cell)
                }

            snapshot["sheets"][sheet]["cellData"][row_idx] = row_data


    return snapshot, col_style
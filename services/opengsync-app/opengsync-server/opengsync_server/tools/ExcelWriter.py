from typing import Literal
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.worksheet.worksheet import Worksheet

BorderStyle = Literal[
    'dashDot','dashDotDot', 'dashed','dotted',
    'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
    'mediumDashed', 'slantDashDot', 'thick', 'thin'
]

class ExcelWriter:
    def __init__(self, dfs: dict[str, pd.DataFrame], index: bool = True):
        self.dfs = dfs
        self.bytes_io = BytesIO()
        self.writer = pd.ExcelWriter(self.bytes_io, engine="openpyxl", mode="w")
        for sheet_name, df in dfs.items():
            df.to_excel(self.writer, sheet_name=sheet_name, index=index)

    def apply_header_style(
        self, sheet_name: str | None, fill_color: str = "E0E0E0",
        font_size: int = 12, bold: bool = True, border: bool = True,
        border_style: BorderStyle = "thin", border_color: str = "000000",
        alignment: Literal["center", "left", "right"] | None = "center"
    ):
        if sheet_name is None:
            for sheet in self.dfs.keys():
                self._apply_header_style(
                    self.writer.sheets[sheet], fill_color, font_size, bold, border, border_style, border_color,
                    alignment=alignment
                )
        else:
            self._apply_header_style(
                self.writer.sheets[sheet_name], fill_color, font_size, bold, border, border_style, border_color, alignment=alignment
            )

    def apply_body_style(
        self, sheet_name: str | None, fill_color: str = "E0E0E0",
        font_size: int = 12, bold: bool = True, border: bool = True,
        border_style: BorderStyle = "thin", border_color: str = "969696",
        alignment: Literal["center", "left", "right"] | None = None
    ):
        if sheet_name is None:
            for sheet in self.dfs.keys():
                self._apply_body_style(
                    self.writer.sheets[sheet], fill_color, font_size, bold, border, border_style, border_color
                )
        else:
            self._apply_body_style(
                self.writer.sheets[sheet_name], fill_color, font_size, bold, border, border_style, border_color
            )
        
        
    def apply_alternating_colors(
        self, sheet_name: str | None, column: str, index: bool = True,
        primary_color: str = "E6F3FF", secondary_color: str = "FFFFFF"
    ):
        """Apply alternating colors based on a column value"""

        if sheet_name is None:
            for sheet in self.dfs.keys():
                self._apply_alternating_colors(
                    self.writer.sheets[sheet], self.dfs[sheet], column, index, primary_color, secondary_color
                )
        else:
            sheet = self.writer.sheets[sheet_name]
            df = self.dfs[sheet_name]
            self._apply_alternating_colors(
                sheet, df, column, index, primary_color, secondary_color
            )

    def save_xlsx(self, path: str):
        with open(path, "wb") as f:
            f.write(self.bytes_io.getvalue())

    def apply_column_width(
        self, sheet_name: str | None, min_width: int = 10, max_width: int = 50, padding: int = 5,
    ):
        if sheet_name is None:
            for sheet in self.dfs.keys():
                self._apply_column_width(
                    self.writer.sheets[sheet], min_width, max_width, padding
                )
        else:
            if sheet_name not in self.dfs:
                print(f"Sheet '{sheet_name}' not found in DataFrames")
                return
            self._apply_column_width(
                self.writer.sheets[sheet_name], min_width, max_width, padding
            )

    @staticmethod
    def _apply_column_width(
        sheet: Worksheet, min_width: int = 10, max_width: int = 50, padding: int = 5,
    ):        
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # type: ignore
            for cell in column:
                cell.alignment = Alignment(wrap_text=False, vertical="top")
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + padding, min_width), max_width)
            sheet.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _apply_body_style(
        sheet: Worksheet, fill_color: str = "E0E0E0",
        font_size: int = 12, bold: bool = True, border: bool = True,
        border_style: BorderStyle = "thin", border_color: str = "000000",
        alignment: Literal["center", "left", "right"] | None = None
    ):
        body_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        body_font = Font(size=font_size, bold=bold)
        body_alignment = Alignment(horizontal="left", vertical=alignment, wrap_text=True)
        border_side = Side(style=border_style, color=border_color)
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = body_fill
                cell.font = body_font
                cell.alignment = body_alignment
                if border:
                    cell.border = thin_border

    @staticmethod
    def _apply_header_style(
        sheet: Worksheet, fill_color: str = "E0E0E0",
        font_size: int = 12, bold: bool = True, border: bool = True,
        border_style: BorderStyle = "thin", border_color: str = "000000",
        alignment: Literal["center", "left", "right"] | None = "center"
    ):
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        header_font = Font(size=font_size, bold=bold)
        header_alignment = Alignment(horizontal=alignment, vertical=alignment, wrap_text=True)
        border_side = Side(style=border_style, color=border_color)
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            if border:
                cell.border = thin_border

    @staticmethod
    def _apply_alternating_colors(
        sheet: Worksheet, df: pd.DataFrame, column: str, index: bool = True,
        primary_color: str = "E6F3FF", secondary_color: str = "FFFFFF"
    ):
        """Apply alternating colors based on a column value"""
        primary_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
        secondary_fill = PatternFill(start_color=secondary_color, end_color=secondary_color, fill_type="solid")
        
        if index:
            col_offset = 2
        else:
            col_offset = 1
        
        col_idx = None
        for col_num, col_name in enumerate(df.columns, col_offset):
            if col_name == column:
                col_idx = col_num
                break
        
        if col_idx is None:
            print(f"Column '{column}' not found in DataFrame")
            return
        
        previous_value = None
        current_fill = primary_fill
        
        for row_num in range(2, len(df) + 2):
            experiment_cell = sheet.cell(row=row_num, column=col_idx)
            current_value = experiment_cell.value

            if current_value != previous_value:
                if previous_value is not None:
                    current_fill = secondary_fill if current_fill == primary_fill else primary_fill
                previous_value = current_value

            for col_num in range(1, len(df.columns) + col_offset):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.fill = current_fill
    
    def __del__(self):
        self.writer.close()
        self.bytes_io.close()

    def get_bytes(self) -> bytes:
        self.writer.close()
        return self.bytes_io.getvalue()
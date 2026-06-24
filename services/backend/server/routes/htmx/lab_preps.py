from typing import Literal
import os
import io

import pandas as pd
from fastapi import APIRouter, Depends, Query, Request, Response
from sqlalchemy import orm
from loguru import logger

from opengsync_db import models, AsyncSession, queries as Q, categories as C, utils

from ...core import dependencies, responses, exceptions as exc
from ...components.tables import HTMXTable, TableCol, StaticSpreadsheet, TextColumn
from ... import forms

router = APIRouter(prefix="/lab_preps", tags=["lab_preps"])


class LabPrepTable(HTMXTable):
    columns = [
        TableCol(title="ID", label="id", col_size=1, searchable=True, sortable=True),
        TableCol(title="Name", label="name", col_size=2, searchable=True, sortable=True),
        TableCol(title="Checklist", label="checklist", col_size=2, choices=C.LabChecklistType.as_selectable(), sortable=True, sort_by="checklist_type_id"),
        TableCol(title="Service", label="service", col_size=2, choices=C.ServiceType.as_selectable(), sortable=True, sort_by="service_type_id"),
        TableCol(title="Status", label="status", col_size=2, choices=C.PrepStatus.as_selectable(), sortable=True, sort_by="status_id"),
        TableCol(title="# Samples", label="num_samples", col_size=1, sortable=True),
        TableCol(title="# Libraries", label="num_libraries", col_size=1, sortable=True),
        TableCol(title="Creator", label="creator", col_size=2, searchable=True),
        TableCol(title="Library Types", label="library_types", col_size=2),
    ]


@router.get("/render-table", name="render_lab_prep_table")
async def render_lab_prep_table(
    status_in: list[C.PrepStatus] | None = Depends(dependencies.parse_enum_ids(enum_type=C.PrepStatus, query_param="status_in")),
    checklist_in: list[C.LabChecklistType] | None = Depends(dependencies.parse_enum_ids(enum_type=C.LabChecklistType, query_param="checklist_in")),
    service_in: list[C.ServiceType] | None = Depends(dependencies.parse_enum_ids(enum_type=C.ServiceType, query_param="service_in")),
    name: str | None = Query(None, description="Optional name to search for"),
    id: int | None = Query(None, description="Optional ID to search for"),
    creator: str | None = Query(None, description="Optional creator name to search for"),
    page: int = Query(0, ge=0, description="Page number, starting from 0"),
    order_by: utils.OrderBy | None = Depends(dependencies.parse_order_by(model=models.LabPrep, default=models.LabPrep.id.desc())),
    session: AsyncSession = Depends(dependencies.db_session),
):
    table = LabPrepTable(route="render_lab_prep_table", page=page, order_by=order_by)
    table.template = "components/tables/lab_prep.html"

    if status_in:
        table.filter_values["status"] = status_in
    if checklist_in:
        table.filter_values["checklist"] = checklist_in
    if service_in:
        table.filter_values["service"] = service_in

    stmt = Q.lab_prep.select(
        status_in=status_in,
        checklist_type_in=checklist_in,
        service_type_in=service_in,
    )

    if name:
        table.active_search_var = "name"
        table.active_query_value = name
    elif id is not None:
        table.active_search_var = "id"
        table.active_query_value = str(id)
    elif creator:
        table.active_search_var = "creator"
        table.active_query_value = creator

    stmt = Q.lab_prep.search(
        name=name,
        creator_name=creator,
        statement=stmt,
    )

    if id is not None and not name and not creator:
        stmt = Q.lab_prep.select(id=id, statement=stmt)

    lab_preps, count = await session.page(
        stmt,
        page=page,
        order_by=order_by,
        options=[
            orm.selectinload(models.LabPrep.creator),
            orm.selectinload(models.LabPrep.libraries),
            orm.with_expression(models.LabPrep._num_samples, models.LabPrep.num_samples.expression),
            orm.with_expression(models.LabPrep._num_libraries, models.LabPrep.num_libraries.expression),
        ],
    )
    table.set_num_pages(count)
    return await table.make_response(lab_preps=lab_preps)


@router.get("/create")
async def render_create_lab_prep_form(
    request: Request,
):
    """Render the create lab prep form."""
    form = forms.models.LabPrepForm(request, form_type="create")
    return await form.make_response()


@router.post("/create")
async def create_lab_prep(response=Depends(forms.models.LabPrepForm.create)) -> Response: return response

@router.get("/{lab_prep_id}/edit")
async def render_edit_lab_prep_form(
    lab_prep_id: int,
    request: Request,
    session: AsyncSession = Depends(dependencies.db_session),
):
    """Render the edit lab prep form."""
    lab_prep = await session.get_one(Q.lab_prep.select(id=lab_prep_id))
    form = forms.models.LabPrepForm(request, form_type="edit", lab_prep=lab_prep)
    return await form.make_response()


@router.post("/{lab_prep_id}/edit")
async def edit_lab_prep(response=Depends(forms.models.LabPrepForm.edit)) -> Response: return response


@router.post("/{lab_prep_id}/uncomplete")
async def uncomplete_lab_prep(
    lab_prep_id: int,
    request: Request,
    session: AsyncSession = Depends(dependencies.db_session),
):
    """Revert a completed lab prep back to preparing."""
    lab_prep = await session.get_one(Q.lab_prep.select(id=lab_prep_id))
    lab_prep.status_id = C.PrepStatus.PREPARING.id

    return await responses.htmx_response(
        redirect=request.url_for("lab_prep", lab_prep_id=lab_prep.id),
        flash=responses.flash("Lab prep reverted to preparing!", "success"),
    )

@router.delete("/{lab_prep_id}/delete")
async def delete_lab_prep(
    lab_prep_id: int,
    request: Request,
    session: AsyncSession = Depends(dependencies.db_session),
):
    """Delete a lab prep (only if still in PREPARING status)."""
    lab_prep = await session.get_one(Q.lab_prep.select(id=lab_prep_id))

    if lab_prep.status_id != C.PrepStatus.PREPARING.id:
        return await responses.htmx_response(
            redirect=request.url_for("lab_prep", lab_prep_id=lab_prep.id),
            flash=responses.flash("Cannot delete completed prep.", "warning"),
        )

    await session.delete(lab_prep)
    return await responses.htmx_response(
        redirect=request.url_for("lab_preps"),
        flash=responses.flash("Lab prep deleted!", "success"),
    )

@router.delete("/{lab_prep_id}/remove-library")
async def remove_library_from_prep(
    lab_prep_id: int,
    request: Request,
    library_id: int = Query(...),
    session: AsyncSession = Depends(dependencies.db_session),
):
    """Remove a library from a lab prep."""
    lab_prep = await session.get_one(Q.lab_prep.select(id=lab_prep_id))

    if lab_prep.status_id != C.PrepStatus.PREPARING.id:
        raise exc.BadRequestException("Cannot remove libraries from a completed prep.")

    library = await session.get_one(Q.library.select(id=library_id))
    library.lab_prep = None

    return await responses.htmx_response(
        redirect=request.url_for("lab_prep", lab_prep_id=lab_prep.id),
        flash=responses.flash("Library removed!", "success"),
    )


@router.get("/{lab_prep_id}/checklist")
async def render_lab_prep_checklist(
    lab_prep_id: int,
    session: AsyncSession = Depends(dependencies.db_session),
):
    lab_prep = await session.get_one(
        Q.lab_prep.select(id=lab_prep_id).options(
            orm.selectinload(models.LabPrep.libraries).selectinload(models.Library.sample_links),
            orm.selectinload(models.LabPrep.libraries).selectinload(models.Library.indices),
        )
    )
    checklist = lab_prep.get_checklist()
    return await responses.htmx_response(
        "components/checklists/lab_prep.html",
        lab_prep=lab_prep, **checklist
    )

@router.get("/{lab_prep_id}/mux-spreadsheet")
async def render_lab_prep_mux_spreadsheet(
    lab_prep_id: int,
    session: AsyncSession = Depends(dependencies.db_session),
):
    df = await session.pd.get_lab_prep_pooling_table(lab_prep_id)

    df = df.sort_values(by=["library_name", "sample_pool", "sample_name"])

    mux_table = {
        "sample_name": [],
        "library_name": [],
        "sample_pool": [],
        "barcode": [],
        "pattern": [],
        "read": []
    }

    for _, row in df.iterrows():
        if row["mux_type_id"] is None:
            continue
        
        mux_table["sample_name"].append(row["sample_name"])
        mux_table["library_name"].append(row["library_name"])
        mux_table["sample_pool"].append(row["sample_pool"])
        if (mux := row.get("mux")) is None:
            mux_table["barcode"].append(None)
            mux_table["pattern"].append(None)
            mux_table["read"].append(None)
        else:
            mux_table["barcode"].append(mux.get("barcode"))
            mux_table["pattern"].append(mux.get("pattern"))
            mux_table["read"].append(mux.get("read"))

    df = pd.DataFrame(mux_table)
    if df["pattern"].isna().all():
        df = df.drop(columns=["pattern"])
    if df["read"].isna().all():
        df = df.drop(columns=["read"])
        
    columns = []
    for col in df.columns:
        columns.append(
            TextColumn(
                col,
                col.replace("_", " ").title().replace("Id", "ID").replace("Cmo", "CMO"),
                {
                    "sample_name": 250,
                    "library_name": 300,
                    "sample_pool": 200,
                    "barcode": 100,
                    "read": 80,
                    "pattern": 150
                }.get(col, 100),
                max_length=1000
            )
        )

    spreadsheet = StaticSpreadsheet(df, columns=columns, id=f"lab_prep_mux_table-{lab_prep_id}")
    return await spreadsheet.render()


@router.get("/{lab_prep_id}/prep-spreadsheet-template")
async def download_lab_prep_spreadsheet_template(
    lab_prep_id: int,
    direction: str = Query("rows", description="Whether to return the template in rows or columns format"),
    checklist: C.LabChecklistType | None = Depends(dependencies.parse_enum_id(enum_type=C.LabChecklistType, query_param="checklist_id")),
    session: AsyncSession = Depends(dependencies.db_session),
):
    import openpyxl
    from openpyxl import styles as openpyxl_styles
    from openpyxl.utils import get_column_letter

    if direction not in ("rows", "columns"):
        raise exc.BadRequestException(f"Invalid direction: {direction}")
    
    lab_prep = await session.get_one(
        Q.lab_prep.select(id=lab_prep_id).options(
            orm.selectinload(models.LabPrep.libraries).selectinload(models.Library.sample_links),
            orm.selectinload(models.LabPrep.libraries).selectinload(models.Library.indices),
            orm.selectinload(models.LabPrep.libraries).selectinload(models.Library.seq_request).selectinload(models.SeqRequest.requestor),
        )
    )

    checklist = checklist or lab_prep.checklist_type

    filepath = os.path.join("/static", "resources", "templates", "library_prep", checklist.prep_file_name)

    if not os.path.exists(filepath):
        logger.error(f"File not found: {filepath}")
        raise FileNotFoundError(f"Template file not found: {filepath}")

    template = openpyxl.load_workbook(filepath)

    n = 12 if direction == "rows" else 8
    pattern = [True] * n + [False] * n

    def if_color(i):
        return pattern[i]

    prep_table = template["prep_table"]
    column_mapping: dict[str, str] = {}
    
    for col_i in range(0, min(prep_table.max_column, 96)):
        col = get_column_letter(col_i + 1)
        column_name = prep_table[f"{col}1"].value
        column_mapping[column_name] = col
        
        for row_idx, cell in enumerate(prep_table[col][1:]):
            if if_color(row_idx % (n * 2)):
                cell.fill = openpyxl_styles.PatternFill(start_color="ced4da", end_color="ced4da", fill_type="solid")
            else:
                cell.fill = openpyxl_styles.PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

    for row_idx, cell in enumerate(prep_table[column_mapping["plate_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")

    for row_idx, cell in enumerate(prep_table[column_mapping["index_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")
        
    for i, library in enumerate(lab_prep.libraries):
        library_id_cell = prep_table[f"{column_mapping['library_id']}{i + 2}"]
        library_name_cell = prep_table[f"{column_mapping['library_name']}{i + 2}"]
        requestor_cell = prep_table[f"{column_mapping['requestor']}{i + 2}"]
        sequence_i7_cell = prep_table[f"{column_mapping['sequence_i7']}{i + 2}"]
        sequence_i5_cell = prep_table[f"{column_mapping['sequence_i5']}{i + 2}"]
        # kit_i7_cell = active_sheet[f"{column_mapping['kit_i7']}{i + 2}"]
        # kit_i5_cell = active_sheet[f"{column_mapping['kit_i5']}{i + 2}"]
        name_i7_cell = prep_table[f"{column_mapping['name_i7']}{i + 2}"]
        name_i5_cell = prep_table[f"{column_mapping['name_i5']}{i + 2}"]
        library_id_cell.value = library.id
        library_name_cell.value = library.name
        requestor_cell.value = library.seq_request.requestor.name
        if len(library.indices) > 0:
            name_i7_cell.value = library.indices[0].name_i7
            name_i5_cell.value = library.indices[0].name_i5
        sequence_i7_cell.value = library.sequences_i7_str(";")
        sequence_i5_cell.value = library.sequences_i5_str(";")

    if "FLEX_table" in template.sheetnames:
        flex_table = template["FLEX_table"]
        column_mapping: dict[str, str] = {}
    
        for col_i in range(0, min(flex_table.max_column, 96)):
            col = get_column_letter(col_i + 1)
            column_name = flex_table[f"{col}1"].value
            column_mapping[column_name] = col

        i = 2
        for library in lab_prep.libraries:
            if library.type == C.LibraryType.TENX_SC_GEX_FLEX:
                for sample_link in library.sample_links:
                    sample_num_cell = flex_table[f"{column_mapping['sample_num']}{i}"]
                    sample_name_cell = flex_table[f"{column_mapping['sample_name']}{i}"]
                    sample_num_cell.value = i - 1
                    sample_name_cell.value = sample_link.sample.name
                    i += 1

    if "10X_table" in template.sheetnames:
        tenx_table = template["10X_table"]
        column_mapping: dict[str, str] = {}
    
        for col_i in range(0, min(tenx_table.max_column, 96)):
            col = get_column_letter(col_i + 1)
            column_name = tenx_table[f"{col}1"].value
            column_mapping[column_name] = col

        i = 2
        for library in lab_prep.libraries:
            if library.type in (C.LibraryType.TENX_SC_GEX_3PRIME, C.LibraryType.TENX_SC_GEX_5PRIME):
                for sample_link in library.sample_links:
                    sample_num_cell = tenx_table[f"{column_mapping['sample_num']}{i}"]
                    sample_name_cell = tenx_table[f"{column_mapping['sample_name']}{i}"]
                    library_name_cell = tenx_table[f"{column_mapping['library_name']}{i}"]
                    library_id_cell = tenx_table[f"{column_mapping['library_id']}{i}"]
                    requestor_cell = tenx_table[f"{column_mapping['requestor']}{i}"]

                    sample_num_cell.value = i - 1
                    sample_name_cell.value = sample_link.sample.name
                    library_name_cell.value = library.name
                    library_id_cell.value = library.id
                    requestor_cell.value = library.seq_request.requestor.name
                    i += 1
            
    bytes_io = io.BytesIO()
    template.save(bytes_io)
    bytes_io.seek(0)

    return await responses.bytes_response(
        bytes_io,
        filename=f"{lab_prep.name}_{lab_prep.checklist_type.abbreviation}_{direction}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/{lab_prep_id}/prep-table-upload")
async def render_lab_prep_table_upload_form(response=Depends(forms.actions.LibraryPrepTableForm.render)): return response

@router.post("/{lab_prep_id}/prep-table-upload", name="lab_prep_table_upload")
async def upload_lab_prep_table(response=Depends(forms.actions.LibraryPrepTableForm.upload)) -> Response: return response


@router.post("/{lab_prep_id}/complete")
async def complete_lab_prep(
    lab_prep_id: int,
    request: Request,
    session: AsyncSession = Depends(dependencies.db_session),
):
    """Mark a lab prep as completed."""
    lab_prep = await session.get_one(
        Q.lab_prep.select(id=lab_prep_id),
        options=[
            orm.selectinload(models.LabPrep.pools),
            orm.selectinload(models.LabPrep.libraries).selectinload(
                models.Library.seq_request
            ).selectinload(models.SeqRequest.libraries),
        ],
    )

    for pool in lab_prep.pools:
        if pool.status < C.PoolStatus.STORED:
            pool.status = C.PoolStatus.STORED

    for library in lab_prep.libraries:
        is_prepared = all(
            sr_lib.status >= C.LibraryStatus.POOLED
            for sr_lib in library.seq_request.libraries
        )
        if is_prepared:
            library.seq_request.status = C.SeqRequestStatus.PREPARED

    lab_prep.status_id = C.PrepStatus.COMPLETED.id

    return await responses.htmx_response(
        redirect=request.url_for("lab_prep", lab_prep_id=lab_prep.id),
        flash=responses.flash("Lab prep completed!", "success"),
    )


@router.get("/{lab_prep_id}/plates")
async def render_lab_prep_plates(
    lab_prep_id: int,
    session: AsyncSession = Depends(dependencies.db_session),
):
    """Render the plates for a lab prep."""
    plates = await session.get_all(
        Q.plate.select(lab_prep_id=lab_prep_id).options(
            orm.selectinload(models.Plate.sample_links)
        ), limit=None
    )
    return await responses.htmx_response(template="components/plates.html", plates=plates)